"""Import the teacher-confirmed three-column Excel roster into YAML."""

from __future__ import annotations

import math
import tempfile
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import yaml

from .config import load_student_roster
from .exceptions import ConfigurationError

EXPECTED_HEADERS = ("学号", "姓名", "github账号名")


def _student_id(value: Any, row: int) -> str:
    if isinstance(value, bool) or value is None:
        raise ConfigurationError(f"Excel 第 {row} 行学号为空或无效")
    if isinstance(value, int):
        result = str(value)
    elif isinstance(value, float) and math.isfinite(value) and value.is_integer():
        result = str(int(value))
    else:
        result = str(value).strip()
    if len(result) != 10 or not result.isdigit():
        raise ConfigurationError(f"Excel 第 {row} 行学号必须是 10 位数字：{result!r}")
    return result


def _required_text(value: Any, row: int, label: str) -> str:
    result = "" if value is None else str(value).strip()
    if not result:
        raise ConfigurationError(f"Excel 第 {row} 行{label}为空")
    return result


def import_students_excel(
    source: str | Path, destination: str | Path, *, force: bool = False
) -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise ConfigurationError(
            "读取 Excel 需要可选依赖，请先执行：python -m pip install 'course-pr-reviewer[excel]'"
        ) from exc

    source_path = Path(source)
    output_path = Path(destination)
    if output_path.exists() and not force:
        raise ConfigurationError(
            f"输出文件已存在，如需覆盖请使用 --force：{output_path}"
        )
    if source_path.suffix.casefold() != ".xlsx":
        raise ConfigurationError("学生名单必须是 .xlsx 文件")
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    except (OSError, ValueError, BadZipFile, InvalidFileException) as exc:
        raise ConfigurationError(f"无法读取 Excel 文件：{source_path}: {exc}") from exc

    try:
        worksheet = workbook.active
        headers = tuple(
            ""
            if worksheet.cell(1, column).value is None
            else str(worksheet.cell(1, column).value).strip()
            for column in range(1, 4)
        )
        if headers != EXPECTED_HEADERS:
            raise ConfigurationError(
                f"Excel 前三列表头必须依次为：{' / '.join(EXPECTED_HEADERS)}；当前为：{' / '.join(headers)}"
            )
        extra_headers = [
            str(worksheet.cell(1, column).value).strip()
            for column in range(4, worksheet.max_column + 1)
            if worksheet.cell(1, column).value not in (None, "")
        ]
        if extra_headers:
            raise ConfigurationError(
                f"Excel 只允许三列，发现额外列：{', '.join(extra_headers)}"
            )
        for row in range(2, worksheet.max_row + 1):
            for column in range(4, worksheet.max_column + 1):
                value = worksheet.cell(row, column).value
                if value is not None and str(value).strip():
                    raise ConfigurationError(
                        f"Excel 只允许三列，第 {row} 行第 {column} 列存在多余内容"
                    )

        students: dict[str, dict[str, Any]] = {}
        github_logins: dict[str, str] = {}
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2
        ):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            student_id = _student_id(values[0], row_number)
            name = _required_text(values[1], row_number, "姓名")
            github = _required_text(values[2], row_number, "GitHub 账号名")
            if student_id in students:
                raise ConfigurationError(f"Excel 存在重复学号：{student_id}")
            login_key = github.casefold()
            if login_key in github_logins:
                raise ConfigurationError(
                    f"Excel 中 GitHub 账号 {github} 同时对应 {github_logins[login_key]} 和 {student_id}"
                )
            students[student_id] = {"name": name, "github": github, "active": True}
            github_logins[login_key] = student_id
    finally:
        workbook.close()

    if not students:
        raise ConfigurationError("Excel 中没有学生记录")

    rendered = yaml.safe_dump(
        {"version": 1, "students": students},
        allow_unicode=True,
        sort_keys=False,
        default_flow_style=False,
    )
    with tempfile.TemporaryDirectory() as directory:
        validation_path = Path(directory) / "students.yml"
        validation_path.write_text(rendered, encoding="utf-8")
        load_student_roster(validation_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(rendered, encoding="utf-8")
    return len(students)
